import fastf1
from fastf1 import get_session
import pandas as pd
import matplotlib.pyplot as plt
import os

# 设置matplotlib字体
plt.rcParams["font.family"] = ["monospace", "Sarasa Term SC Nerd Font"]

import numpy as np
import openpyxl

# 设置matplotlib字体和字号
plt.rcParams["font.family"] = ["monospace", "Sarasa Term SC Nerd Font"]
plt.rcParams["font.size"] = 20

# 启用FastF1缓存
fastf1.Cache.enable_cache("f1_cache")
fastf1.set_log_level("ERROR")


# 直线放缩函数
def linear_scaling(r_value, base_lap, r_max, min_lap, max_lap):
    """直线放缩函数"""
    return np.clip(base_lap * (1 + 0.595 * (1 - r_value / r_max)), min_lap, max_lap)


# 非线性放缩函数
def nonlinear_scaling(r_value, base_lap, r_max, min_lap, max_lap):
    """非线性放缩函数"""
    return np.clip(
        base_lap
        * (1 + 0.65 * (1 - r_value / r_max) * np.exp(-2.5 * (1 - r_value / r_max))),
        min_lap,
        max_lap,
    )


# 获取排位赛最快和最慢圈速
def get_lap_limits(year, gp_name):
    """提取指定赛季和分站排位赛的最快和最慢圈速"""
    try:
        session = get_session(year, gp_name, "Q")
        session.load()
        laps = session.laps

        # 筛选排位赛有效圈速
        qualifying_laps = laps[
            (laps["IsPersonalBest"] == True)
            & (laps["PitOutTime"].isna())
            & (laps["PitInTime"].isna())
        ]

        fastest_lap = qualifying_laps["LapTime"].min().total_seconds()
        slowest_lap = qualifying_laps["LapTime"].max().total_seconds()

        return fastest_lap, slowest_lap
    except Exception as e:
        print(f"获取数据时出错: {e}")
        return None, None


# 从Excel读取西班牙站车手数据
def read_spain_driver_data(file_path):
    """从Excel文件中读取西班牙站车手数据"""
    # 创建车手全名到三字母简称的映射
    name_to_code = {
        "维斯塔潘": "VER",
        "汉密尔顿": "HAM",
        "勒克莱尔": "LEC",
        "佩雷兹": "PER",
        "拉塞尔": "RUS",
        "塞恩斯": "SAI",
        "加斯利": "GAS",
        "角田裕毅": "TSU",
        "阿隆索": "ALO",
        "维特尔": "VET",
        "奥康": "OCO",
        "斯特罗尔": "STR",
        "诺里斯": "NOR",
        "博塔斯": "BOT",
        "皮亚斯特里": "RIC",
        "周冠宇": "ZHO",
        "马格努森": "MAG",
        "阿尔本": "ALB",
        "米克": "MSC",
        "拉提菲": "LAT",
    }

    # 读取Excel文件
    wb = openpyxl.load_workbook(file_path)

    # 获取所有工作表名称
    sheet_names = wb.sheetnames
    print(f"Excel文件中的工作表: {sheet_names}")

    # 尝试找到包含"西班牙"的工作表
    sheet_name = None
    for name in sheet_names:
        if "西班牙" in name or "Spain" in name:
            sheet_name = name
            break

    # 如果没有找到，使用第一个工作表
    if not sheet_name:
        sheet_name = sheet_names[0]

    print(f"使用工作表: {sheet_name}")
    sheet = wb[sheet_name]

    driver_data = {}

    # 遍历行，跳过标题行
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 检查行是否足够长
        if len(row) < 5:
            continue

        driver_name = row[0]  # 车手全名
        r_value = row[4]  # RO列（R值）

        # 只处理车手名非空的行
        if driver_name and driver_name in name_to_code:
            driver_code = name_to_code[driver_name]
            driver_data[driver_code] = r_value

    return driver_data


# 主程序
def main():
    # 设置参数
    year = 2022
    gp_name = "Spain"
    excel_file = os.path.join("tables", gp_name + ".xlsx")

    # 1. 获取最快和最慢圈速
    fastest_lap_seconds, slowest_lap_seconds = get_lap_limits(year, gp_name)
    if fastest_lap_seconds is None or slowest_lap_seconds is None:
        print("未能获取圈速数据")
        return

    print(f"2022赛季{gp_name}站排位赛:")
    print(f"  最快圈速: {fastest_lap_seconds:.3f}秒")
    print(f"  最慢圈速: {slowest_lap_seconds:.3f}秒")

    # 2. 从Excel读取车手数据
    driver_r_values = read_spain_driver_data(excel_file)
    print(f"从Excel读取到{len(driver_r_values)}位车手的R值数据")

    if not driver_r_values:
        print("错误: 未能读取到任何车手数据，请检查Excel文件格式")
        return

    # 计算R最大值
    r_max = max(driver_r_values.values())
    print(f"最大R值: {r_max:.3f}")

    # 3. 获取实际排位赛数据
    session = get_session(year, gp_name, "Q")
    session.load()
    laps = session.laps

    # 筛选排位赛有效圈速
    qualifying_laps = laps[
        (laps["IsPersonalBest"] == True)
        & (laps["PitOutTime"].isna())
        & (laps["PitInTime"].isna())
    ]

    # 按车手分组获取最快圈速
    driver_fastest_laps = qualifying_laps.groupby("Driver")["LapTime"].min()

    # 4. 构建R值网格
    r_grid = np.linspace(280, 310, 100)  # 使用线性网格，便于观察曲线形状
    print(f"构建R值网格: 从{r_grid[0]:.3f}到{r_grid[-1]:.3f}，共{len(r_grid)}个点")

    # 5. 计算两种放缩方法的曲线
    linear_curve = linear_scaling(
        r_grid, fastest_lap_seconds, r_max, fastest_lap_seconds, slowest_lap_seconds
    )
    nonlinear_curve = nonlinear_scaling(
        r_grid, fastest_lap_seconds, r_max, fastest_lap_seconds, slowest_lap_seconds
    )

    # 6. 计算车手点的模拟圈速
    driver_data = []
    for driver_code, r_value in driver_r_values.items():
        # 获取实际圈速
        if driver_code in driver_fastest_laps:
            actual_lap = driver_fastest_laps[driver_code].total_seconds()
        else:
            actual_lap = np.nan

        # 计算两种放缩方法的模拟圈速
        linear_sim_lap = linear_scaling(
            r_value,
            fastest_lap_seconds,
            r_max,
            fastest_lap_seconds,
            slowest_lap_seconds,
        )
        nonlinear_sim_lap = nonlinear_scaling(
            r_value,
            fastest_lap_seconds,
            r_max,
            fastest_lap_seconds,
            slowest_lap_seconds,
        )

        # 获取车手姓名
        driver_name = session.get_driver(driver_code).Abbreviation

        driver_data.append(
            {
                "Driver": driver_name,
                "DriverCode": driver_code,
                "R_Value": r_value,
                "Actual_Lap": actual_lap,
                "Linear_Sim_Lap": linear_sim_lap,
                "Nonlinear_Sim_Lap": nonlinear_sim_lap,
            }
        )

    # 转换为DataFrame
    df = pd.DataFrame(driver_data)
    df = df.sort_values("R_Value", ascending=False)  # 按R值从大到小排序

    # 7. 绘制对比图

    # 图1: t-R曲线对比
    plt.figure(figsize=(12, 8))

    # 绘制两种放缩方法的曲线
    plt.plot(r_grid, linear_curve, "b-", linewidth=2, label="直线放缩 (α=0.595)")
    plt.plot(
        r_grid, nonlinear_curve, "r-", linewidth=2, label="非线性放缩 (α=0.65, γ=2.5)"
    )

    # 添加圈速范围参考线
    plt.axhline(
        y=fastest_lap_seconds,
        color="green",
        linestyle="--",
        alpha=0.5,
        label="最快圈速",
    )
    plt.axhline(
        y=slowest_lap_seconds,
        color="orange",
        linestyle="--",
        alpha=0.5,
        label="最慢圈速",
    )

    # 标注车手点（直线放缩）
    plt.scatter(
        df["R_Value"],
        df["Linear_Sim_Lap"],
        c="blue",
        s=100,
        edgecolors="black",
        linewidth=1,
        marker="o",
        label="车手点（直线放缩）",
        alpha=0.7,
    )

    # 标注车手点（非线性放缩）
    plt.scatter(
        df["R_Value"],
        df["Nonlinear_Sim_Lap"],
        c="red",
        s=100,
        edgecolors="black",
        linewidth=1,
        marker="s",
        label="车手点（非线性放缩）",
        alpha=0.7,
    )

    # 添加车手标签（只标注部分车手，避免过于拥挤）
    for i, row in df.iterrows():
        if i % 3 == 0:  # 每隔3个车手标注一个
            plt.annotate(
                row["Driver"],
                (row["R_Value"], row["Linear_Sim_Lap"]),
                textcoords="offset points",
                xytext=(5, 5),
                ha="left",
                fontsize=10,
                color="blue",
            )
            plt.annotate(
                row["Driver"],
                (row["R_Value"], row["Nonlinear_Sim_Lap"]),
                textcoords="offset points",
                xytext=(5, -15),
                ha="left",
                fontsize=10,
                color="red",
            )

    # 设置轴标签和标题
    plt.xlabel("R值")
    plt.ylabel("模拟圈速 (秒)")
    plt.title("两种放缩方法的t-R曲线对比")
    plt.legend()
    plt.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig("scaling_comparison_tr_curve.png", dpi=300, bbox_inches="tight")
    plt.show()

    # 图2: 真实-模拟值对比
    plt.figure(figsize=(10, 8))

    # 绘制理想线 (y=x)
    min_val = min(
        df["Linear_Sim_Lap"].min(),
        df["Nonlinear_Sim_Lap"].min(),
        df["Actual_Lap"].min(),
    )
    max_val = max(
        df["Linear_Sim_Lap"].max(),
        df["Nonlinear_Sim_Lap"].max(),
        df["Actual_Lap"].max(),
    )
    plt.plot(
        [min_val, max_val], [min_val, max_val], "k--", alpha=0.5, label="理想线 (y=x)"
    )

    # 绘制直线放缩的点
    plt.scatter(
        df["Linear_Sim_Lap"],
        df["Actual_Lap"],
        c="blue",
        s=100,
        edgecolors="black",
        linewidth=1,
        marker="o",
        label="直线放缩",
        alpha=0.7,
    )

    # 绘制非线性放缩的点
    plt.scatter(
        df["Nonlinear_Sim_Lap"],
        df["Actual_Lap"],
        c="red",
        s=100,
        edgecolors="black",
        linewidth=1,
        marker="s",
        label="非线性放缩",
        alpha=0.7,
    )

    # 添加车手标签（只标注部分车手）
    for i, row in df.iterrows():
        if not np.isnan(row["Actual_Lap"]) and i % 3 == 0:
            plt.annotate(
                row["Driver"],
                (row["Linear_Sim_Lap"], row["Actual_Lap"]),
                textcoords="offset points",
                xytext=(5, 5),
                ha="left",
                fontsize=10,
                color="blue",
            )
            plt.annotate(
                row["Driver"],
                (row["Nonlinear_Sim_Lap"], row["Actual_Lap"]),
                textcoords="offset points",
                xytext=(5, -15),
                ha="left",
                fontsize=10,
                color="red",
            )

    # 设置轴标签和标题
    plt.xlabel("模拟圈速 (秒)")
    plt.ylabel("实际圈速 (秒)")
    plt.title("实际圈速 vs 模拟圈速对比")
    plt.legend()
    plt.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig(
        "scaling_comparison_actual_vs_simulated.png", dpi=300, bbox_inches="tight"
    )
    plt.show()

    # 8. 计算并打印统计信息
    print("\n统计信息:")

    # 计算MSE
    linear_mse = np.mean((df["Linear_Sim_Lap"] - df["Actual_Lap"]) ** 2)
    nonlinear_mse = np.mean((df["Nonlinear_Sim_Lap"] - df["Actual_Lap"]) ** 2)

    print(f"直线放缩MSE: {linear_mse:.6f}秒²")
    print(f"非线性放缩MSE: {nonlinear_mse:.6f}秒²")
    print(f"MSE改进: {((linear_mse - nonlinear_mse) / linear_mse * 100):.2f}%")

    # 计算平均绝对误差
    linear_mae = np.mean(np.abs(df["Linear_Sim_Lap"] - df["Actual_Lap"]))
    nonlinear_mae = np.mean(np.abs(df["Nonlinear_Sim_Lap"] - df["Actual_Lap"]))

    print(f"\n直线放缩MAE: {linear_mae:.6f}秒")
    print(f"非线性放缩MAE: {nonlinear_mae:.6f}秒")
    print(f"MAE改进: {((linear_mae - nonlinear_mae) / linear_mae * 100):.2f}%")

    # 计算R_max附近的导数变化
    dr = r_grid[1] - r_grid[0]
    linear_derivatives = np.gradient(linear_curve, dr)
    nonlinear_derivatives = np.gradient(nonlinear_curve, dr)

    # 取R_max附近最后10个点的导数平均值
    linear_near_rmax = np.mean(np.abs(linear_derivatives[-10:]))
    nonlinear_near_rmax = np.mean(np.abs(nonlinear_derivatives[-10:]))

    print(f"\nR_max附近导数绝对值:")
    print(f"直线放缩: {linear_near_rmax:.6f}")
    print(f"非线性放缩: {nonlinear_near_rmax:.6f}")
    print(
        f"导数减小: {((linear_near_rmax - nonlinear_near_rmax) / linear_near_rmax * 100):.2f}%"
    )

    # 输出放缩函数表达式
    print("\n放缩函数表达式:")
    print("直线放缩: t = base_lap * (1 + 0.595 * (1 - R/R_max))")
    print(
        "非线性放缩: t = base_lap * (1 + 0.65 * (1 - R/R_max) * exp(-2.5*(1 - R/R_max)))"
    )


if __name__ == "__main__":
    main()
