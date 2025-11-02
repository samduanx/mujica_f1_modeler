import fastf1
from fastf1 import get_session
import pandas as pd
import matplotlib.pyplot as plt

n  # 设置matplotlib字体
plt.rcParams["font.family"] = ["monospace", "Sarasa Term SC Nerd Font"]

import numpy as np
from scipy.interpolate import interp1d
from scipy.stats import linregress
import openpyxl

# 设置matplotlib字体和字号
plt.rcParams["font.family"] = ["monospace", "Sarasa Term SC Nerd Font"]
plt.rcParams["font.size"] = 20

# 启用FastF1缓存
fastf1.Cache.enable_cache("f1_cache")
fastf1.set_log_level("ERROR")


# 圈速计算函数
def calculate_lap_time(r_value, base_lap, r_max, alpha, min_lap, max_lap):
    """计算模拟圈速的函数，确保在最快和最慢圈速之间"""
    # 计算原始模拟圈速
    simulated_lap = base_lap * (1 + alpha * (1 - r_value / r_max))

    # 确保圈速在指定范围内
    simulated_lap = np.clip(simulated_lap, min_lap, max_lap)

    return simulated_lap


# 获取排位赛最快和最慢圈速
def get_lap_limits(year, gp_name):
    """提取指定赛季和分站排位赛的最快和最慢圈速"""
    try:
        session = get_session(year, gp_name, "Q")
        session.load()
        laps = session.laps

        # 筛选排位赛有效圈速
        qualifying_laps = laps[
            (laps["IsPersonalBest"] == True)
            & (laps["PitOutTime"].isna())
            & (laps["PitInTime"].isna())
        ]

        fastest_lap = qualifying_laps["LapTime"].min().total_seconds()
        slowest_lap = qualifying_laps["LapTime"].max().total_seconds()

        return fastest_lap, slowest_lap
    except Exception as e:
        print(f"获取数据时出错: {e}")
        return None, None


# 从Excel读取西班牙站车手数据
def read_spain_driver_data(file_path):
    """从Excel文件中读取西班牙站车手数据"""
    # 创建车手全名到三字母简称的映射
    name_to_code = {
        "维斯塔潘": "VER",
        "汉密尔顿": "HAM",
        "勒克莱尔": "LEC",
        "佩雷兹": "PER",
        "拉塞尔": "RUS",
        "塞恩斯": "SAI",
        "加斯利": "GAS",
        "角田裕毅": "TSU",
        "阿隆索": "ALO",
        "维特尔": "VET",
        "奥康": "OCO",
        "斯特罗尔": "STR",
        "诺里斯": "NOR",
        "博塔斯": "BOT",
        "皮亚斯特里": "RIC",
        "周冠宇": "ZHO",
        "马格努森": "MAG",
        "阿尔本": "ALB",
        "米克": "MSC",
        "拉提菲": "LAT",
    }

    # 读取Excel文件
    wb = openpyxl.load_workbook(file_path)

    # 获取所有工作表名称
    sheet_names = wb.sheetnames
    print(f"Excel文件中的工作表: {sheet_names}")

    # 尝试找到包含"西班牙"的工作表
    sheet_name = None
    for name in sheet_names:
        if "西班牙" in name or "Spain" in name:
            sheet_name = name
            break

    # 如果没有找到，使用第一个工作表
    if not sheet_name:
        sheet_name = sheet_names[0]

    print(f"使用工作表: {sheet_name}")
    sheet = wb[sheet_name]

    driver_data = {}

    # 打印前几行数据用于调试
    print("\nExcel文件前5行数据:")
    for i, row in enumerate(sheet.iter_rows(max_row=5, values_only=True)):
        print(f"行{i + 1}: {row}")

    # 遍历行，跳过标题行
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 检查行是否足够长
        if len(row) < 5:
            continue

        driver_name = row[0]  # 车手全名
        r_value = row[4]  # RO列（R值）

        # 只处理车手名非空的行
        if driver_name and driver_name in name_to_code:
            driver_code = name_to_code[driver_name]
            driver_data[driver_code] = r_value
            print(f"添加车手: {driver_name} ({driver_code}) -> R值: {r_value}")

    return driver_data


# 评估alpha值的函数
def evaluate_alpha(
    alpha, r_grid, fastest_lap_seconds, r_max, driver_r_values, min_lap, max_lap
):
    """评估alpha值的综合得分"""
    # 计算网格点圈速
    simulated_lap_grid = calculate_lap_time(
        r_grid, fastest_lap_seconds, r_max, alpha, min_lap, max_lap
    )

    # 创建插值函数
    interp_func = interp1d(
        r_grid, simulated_lap_grid, kind="linear", fill_value="extrapolate"
    )

    # 计算每个车手的模拟圈速和曲线圈速
    simulated_laps = []
    curve_laps = []
    r_values = []

    for driver_code, r_value in driver_r_values.items():
        simulated_lap = calculate_lap_time(
            r_value, fastest_lap_seconds, r_max, alpha, min_lap, max_lap
        )
        curve_lap = interp_func(r_value)

        simulated_laps.append(simulated_lap)
        curve_laps.append(curve_lap)
        r_values.append(r_value)

    # 计算MSE
    mse = np.mean((np.array(simulated_laps) - np.array(curve_laps)) ** 2)

    # 计算y=x曲线上下的点数
    above_count = sum(1 for s, c in zip(simulated_laps, curve_laps) if s > c)
    below_count = sum(1 for s, c in zip(simulated_laps, curve_laps) if s < c)
    on_count = sum(1 for s, c in zip(simulated_laps, curve_laps) if s == c)

    # 计算点数分布平衡得分（0-1，1表示完全平衡）
    total_points = above_count + below_count
    if total_points == 0:
        balance_score = 0
    else:
        balance_score = 1 - abs(above_count - below_count) / total_points

    # 检查y=x曲线下是否有点
    has_below_points = below_count > 0

    # 计算点数分布均匀性（上下点数比例接近1:1）
    ideal_ratio = 0.5
    if total_points > 0:
        above_ratio = above_count / total_points
        below_ratio = below_count / total_points
        uniformity_score = 1 - (
            abs(above_ratio - ideal_ratio) + abs(below_ratio - ideal_ratio)
        )
    else:
        uniformity_score = 0

    # 综合得分（加权平均）
    # 权重：MSE (0.7), 均匀性 (0.3)
    mse_score = 1 / (1 + mse)  # 将MSE转换为得分（越小越好）

    composite_score = 0.7 * mse_score + 0.3 * uniformity_score

    return {
        "alpha": alpha,
        "mse": mse,
        "balance_score": balance_score,
        "uniformity_score": uniformity_score,
        "has_below_points": has_below_points,
        "above_count": above_count,
        "below_count": below_count,
        "composite_score": composite_score,
        "simulated_laps": simulated_laps,
        "curve_laps": curve_laps,
        "r_values": r_values,
    }


# 主程序
def main():
    # 设置参数
    year = 2022
    gp_name = "Spain"
    excel_file = os.path.join("tables", gp_name + ".xlsx")

    # 1. 获取最快和最慢圈速
    fastest_lap_seconds, slowest_lap_seconds = get_lap_limits(year, gp_name)
    if fastest_lap_seconds is None or slowest_lap_seconds is None:
        print("未能获取圈速数据")
        return

    print(f"2022赛季{gp_name}站排位赛:")
    print(f"  最快圈速: {fastest_lap_seconds:.3f}秒")
    print(f"  最慢圈速: {slowest_lap_seconds:.3f}秒")
    print(f"  圈速范围: {slowest_lap_seconds - fastest_lap_seconds:.3f}秒")

    # 2. 从Excel读取车手数据
    driver_r_values = read_spain_driver_data(excel_file)
    print(f"从Excel读取到{len(driver_r_values)}位车手的R值数据")

    if not driver_r_values:
        print("错误: 未能读取到任何车手数据，请检查Excel文件格式")
        return

    # 计算R最大值
    r_max = max(driver_r_values.values())
    print(f"最大R值: {r_max:.3f}")

    # 3. 获取实际排位赛数据
    session = get_session(year, gp_name, "Q")
    session.load()
    laps = session.laps

    # 筛选排位赛有效圈速
    qualifying_laps = laps[
        (laps["IsPersonalBest"] == True)
        & (laps["PitOutTime"].isna())
        & (laps["PitInTime"].isna())
    ]

    # 按车手分组获取最快圈速
    driver_fastest_laps = qualifying_laps.groupby("Driver")["LapTime"].min()

    # 4. 构建对数分布的R值网格
    r_grid = np.logspace(np.log10(310), np.log10(280), num=50)
    print(f"构建对数R值网格: 从{r_grid[0]:.3f}到{r_grid[-1]:.3f}，共{len(r_grid)}个点")

    # 5. 扫描alpha参数，寻找最优值
    alpha_values = np.arange(0.595, 0.615, 0.001)  # 从0.05到1.5，步长0.05
    evaluation_results = []

    print("\n开始alpha参数优化...")
    for alpha in alpha_values:
        result = evaluate_alpha(
            alpha,
            r_grid,
            fastest_lap_seconds,
            r_max,
            driver_r_values,
            fastest_lap_seconds,
            slowest_lap_seconds,
        )
        evaluation_results.append(result)
        print(
            f"alpha={alpha:.2f}, MSE={result['mse']:.6f}, "
            f"均匀性={result['uniformity_score']:.3f}, "
            f"综合得分={result['composite_score']:.3f}, "
            f"上/下点数={result['above_count']}/{result['below_count']}"
        )

    # 找到最优alpha（综合得分最高）
    optimal_result = max(evaluation_results, key=lambda x: x["composite_score"])
    optimal_alpha = optimal_result["alpha"]

    print(f"\n最优alpha值: {optimal_alpha:.2f}")
    print(f"最优参数详情:")
    print(f"  MSE: {optimal_result['mse']:.6f}")
    print(f"  均匀性得分: {optimal_result['uniformity_score']:.3f}")
    print(f"  是否有下方点: {optimal_result['has_below_points']}")
    print(
        f"  上/下点数: {optimal_result['above_count']}/{optimal_result['below_count']}"
    )
    print(f"  综合得分: {optimal_result['composite_score']:.3f}")

    # 6. 使用最优alpha计算最终结果
    # 计算车手模拟圈速
    results = []
    for driver_code, r_value in driver_r_values.items():
        simulated_lap = calculate_lap_time(
            r_value,
            fastest_lap_seconds,
            r_max,
            optimal_alpha,
            fastest_lap_seconds,
            slowest_lap_seconds,
        )

        # 获取实际圈速
        if driver_code in driver_fastest_laps:
            actual_lap = driver_fastest_laps[driver_code].total_seconds()
        else:
            actual_lap = np.nan

        # 获取车手姓名
        driver_name = session.get_driver(driver_code).Abbreviation

        results.append(
            {
                "Driver": driver_name,
                "DriverCode": driver_code,
                "R_Value": r_value,
                "Actual_Lap": actual_lap,
                "Simulated_Lap": simulated_lap,
                "Difference": actual_lap - simulated_lap
                if not np.isnan(actual_lap)
                else np.nan,
            }
        )

    # 转换为DataFrame
    df = pd.DataFrame(results)
    df = df.sort_values("Simulated_Lap")

    # 计算网格点圈速
    simulated_lap_grid = calculate_lap_time(
        r_grid,
        fastest_lap_seconds,
        r_max,
        optimal_alpha,
        fastest_lap_seconds,
        slowest_lap_seconds,
    )

    # 创建插值函数
    interp_func = interp1d(
        r_grid, simulated_lap_grid, kind="linear", fill_value="extrapolate"
    )

    # 计算每个车手在曲线上的预测圈速
    df["Curve_Lap"] = interp_func(df["R_Value"])

    # 计算MSE
    mse = np.mean((df["Simulated_Lap"] - df["Curve_Lap"]) ** 2)
    print(f"使用最优alpha的MSE: {mse:.6f}秒²")

    # 7. 绘制四个独立的图表

    # 图1: 车手R值与模拟圈速对比
    plt.figure(figsize=(14, 8))
    ax1 = plt.gca()
    ax2 = ax1.twinx()

    bar_width = 0.35
    x = np.arange(len(df))

    # R值条形图
    bars1 = ax1.bar(
        x - bar_width / 2, df["R_Value"], bar_width, color="skyblue", label="R值"
    )

    # 模拟圈速条形图
    bars2 = ax2.bar(
        x + bar_width / 2,
        df["Simulated_Lap"],
        bar_width,
        color="salmon",
        label="模拟圈速",
    )

    # 添加圈速范围参考线
    ax2.axhline(
        y=fastest_lap_seconds,
        color="green",
        linestyle="--",
        alpha=0.5,
        label="最快圈速",
    )
    ax2.axhline(
        y=slowest_lap_seconds, color="red", linestyle="--", alpha=0.5, label="最慢圈速"
    )

    # 设置轴标签和标题
    ax1.set_xlabel("车手")
    ax1.set_ylabel("R值", color="skyblue")
    ax2.set_ylabel("模拟圈速 (秒)", color="salmon")
    plt.title(f"车手R值与模拟圈速对比 (西班牙站, α={optimal_alpha:.2f})")
    ax1.set_xticks(x)
    ax1.set_xticklabels(df["Driver"], rotation=45, ha="right")

    # 添加图例
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper right")

    # 添加数值标签
    for bar in bars1:
        height = bar.get_height()
        ax1.text(
            bar.get_x() + bar.get_width() / 2,
            height,
            f"{height:.1f}",
            ha="center",
            va="bottom",
            fontsize=12,
        )

    for bar in bars2:
        height = bar.get_height()
        ax2.text(
            bar.get_x() + bar.get_width() / 2,
            height,
            f"{height:.3f}",
            ha="center",
            va="bottom",
            fontsize=12,
        )

    plt.tight_layout()
    plt.savefig("driver_r_vs_laptime.png", dpi=300, bbox_inches="tight")
    plt.show()

    # 图2: t-R曲线和车手点分布
    plt.figure(figsize=(12, 8))
    # 绘制t-R曲线
    plt.plot(
        r_grid,
        simulated_lap_grid,
        "b-",
        linewidth=2,
        label=f"t-R曲线 (α={optimal_alpha:.2f})",
    )

    # 添加圈速范围参考线
    plt.axhline(
        y=fastest_lap_seconds,
        color="green",
        linestyle="--",
        alpha=0.5,
        label="最快圈速",
    )
    plt.axhline(
        y=slowest_lap_seconds, color="red", linestyle="--", alpha=0.5, label="最慢圈速"
    )

    # 标注车手点
    scatter = plt.scatter(
        df["R_Value"],
        df["Simulated_Lap"],
        c=df["R_Value"],
        cmap="viridis",
        s=100,
        edgecolors="black",
        linewidth=1,
        label="车手点",
    )

    # 添加车手标签
    for i, row in df.iterrows():
        plt.annotate(
            row["Driver"],
            (row["R_Value"], row["Simulated_Lap"]),
            textcoords="offset points",
            xytext=(5, 5),
            ha="left",
            fontsize=10,
        )

    # 设置轴标签和标题
    plt.xlabel("R值")
    plt.ylabel("模拟圈速 (秒)")
    plt.title(f"t-R曲线与车手点分布 (西班牙站, α={optimal_alpha:.2f})")
    plt.legend()
    plt.grid(True, alpha=0.3)

    # 添加颜色条
    cbar = plt.colorbar(scatter)
    cbar.set_label("R值")

    plt.tight_layout()
    plt.savefig("tr_curve.png", dpi=300, bbox_inches="tight")
    plt.show()

    # 图3: 实际圈速与模拟圈速对比
    plt.figure(figsize=(10, 8))
    plt.scatter(df["Simulated_Lap"], df["Actual_Lap"], color="blue", alpha=0.7)

    # 添加理想线 (y=x)
    min_val = min(df["Simulated_Lap"].min(), df["Actual_Lap"].min())
    max_val = max(df["Simulated_Lap"].max(), df["Actual_Lap"].max())
    plt.plot(
        [min_val, max_val], [min_val, max_val], "r--", alpha=0.5, label="理想线 (y=x)"
    )

    # 添加车手标签
    for i, row in df.iterrows():
        if not np.isnan(row["Actual_Lap"]):
            plt.annotate(
                row["Driver"],
                (row["Simulated_Lap"], row["Actual_Lap"]),
                textcoords="offset points",
                xytext=(5, 5),
                ha="left",
                fontsize=10,
            )

    # 设置轴标签和标题
    plt.xlabel("模拟圈速 (秒)")
    plt.ylabel("实际圈速 (秒)")
    plt.title(f"实际圈速 vs 模拟圈速 (西班牙站, α={optimal_alpha:.2f})")
    plt.legend()
    plt.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.savefig("actual_vs_simulated.png", dpi=300, bbox_inches="tight")
    plt.show()

    # 图4: 车手点与曲线点的误差分布
    plt.figure(figsize=(12, 6))
    errors = df["Simulated_Lap"] - df["Curve_Lap"]
    plt.bar(df["Driver"], errors, color="green", alpha=0.7)
    plt.axhline(y=0, color="red", linestyle="--", alpha=0.5)
    plt.xlabel("车手")
    plt.ylabel("误差 (秒)")
    plt.title("车手点与曲线点的误差分布")
    plt.xticks(rotation=45, ha="right")
    plt.grid(True, alpha=0.3)

    # 添加误差值标签
    for i, error in enumerate(errors):
        plt.text(
            i,
            error + 0.001 if error >= 0 else error - 0.001,
            f"{error:.4f}",
            ha="center",
            va="bottom" if error >= 0 else "top",
            fontsize=10,
        )

    plt.tight_layout()
    plt.savefig("error_distribution.png", dpi=300, bbox_inches="tight")
    plt.show()

    # 打印统计信息
    print("\n统计信息:")
    print(f"平均差异: {df['Difference'].mean():.3f}秒")
    print(f"最大差异: {df['Difference'].abs().max():.3f}秒")
    print(f"最小差异: {df['Difference'].abs().min():.3f}秒")
    print(f"差异标准差: {df['Difference'].std():.3f}秒")

    # 找出最接近和最偏离理想的车手
    closest = df.loc[df["Difference"].abs().idxmin()]
    furthest = df.loc[df["Difference"].abs().idxmax()]

    print(
        f"\n最接近理想的车手: {closest['Driver']} (差异: {closest['Difference']:.3f}秒)"
    )
    print(
        f"最偏离理想的车手: {furthest['Driver']} (差异: {furthest['Difference']:.3f}秒)"
    )


if __name__ == "__main__":
    main()
